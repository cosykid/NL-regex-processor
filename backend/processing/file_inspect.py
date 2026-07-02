"""Lightweight upload inspection.

Returns the column names and a small preview (first N rows) so the UI can let
the user pick target columns. Critically, this reads only the *first rows* of
the file — never the whole thing — so even a multi-GB upload is inspected
cheaply. The heavy, full-file work happens later in Spark.

Every entry point takes a *source* that is either a local filesystem path (as
the tests and the local backend use) or an already-open **seekable binary
stream** (as the S3 backend provides via ``storage.open_binary``). That keeps
the byte-offset windowing logic identical whether the bytes live on disk or in
object storage.
"""
from __future__ import annotations

import csv
from contextlib import contextmanager

from django.conf import settings

from jobs.models import UploadedFile


@contextmanager
def _as_binary(source):
    """Yield a binary stream for ``source``.

    A path-like is opened (and closed) here; an already-open stream is yielded
    as-is and left open for its owner to close.
    """
    if hasattr(source, "read"):
        yield source
    else:
        fh = open(source, "rb")
        try:
            yield fh
        finally:
            fh.close()


def detect_kind(filename: str) -> str:
    lower = filename.lower()
    if lower.endswith((".xlsx", ".xlsm", ".xls")):
        return UploadedFile.Kind.EXCEL
    return UploadedFile.Kind.CSV


def inspect(source, kind: str, preview_rows: int | None = None) -> dict:
    if preview_rows is None:
        preview_rows = settings.UPLOAD_PREVIEW_ROWS
    if kind == UploadedFile.Kind.EXCEL:
        return _inspect_excel(source, preview_rows)
    return _inspect_csv(source, preview_rows)


def _inspect_csv(source, preview_rows: int) -> dict:
    rows: list[dict] = []
    with _as_binary(source) as fh:
        reader = csv.reader(_decoded_lines(fh))
        try:
            columns = next(reader)
        except StopIteration:
            return {"columns": [], "preview_rows": []}
        columns = [c.strip() for c in columns]
        if columns:  # strip a UTF-8 BOM the first header cell may carry
            columns[0] = columns[0].lstrip("﻿")
        for i, raw in enumerate(reader):
            if i >= preview_rows:
                break
            rows.append(_row_to_dict(columns, raw))
    return {"columns": columns, "preview_rows": rows}


def _inspect_excel(source, preview_rows: int) -> dict:
    # read_only mode streams rows lazily instead of loading the workbook.
    from openpyxl import load_workbook

    with _as_binary(source) as fh:
        wb = load_workbook(filename=fh, read_only=True, data_only=True)
        try:
            ws = wb.active
            columns: list[str] = []
            rows: list[dict] = []
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i == 0:
                    columns = [str(c).strip() if c is not None else "" for c in row]
                    continue
                if i > preview_rows:
                    break
                values = ["" if v is None else str(v) for v in row]
                rows.append(_row_to_dict(columns, values))
            return {"columns": columns, "preview_rows": rows}
        finally:
            wb.close()


def read_window(
    source, kind: str, columns: list[str], cursor: str | None, limit: int
) -> dict:
    """Read the next ``limit``-row window of the raw upload.

    Powers browsing the *original* file (before any transformation) as the UI
    lazily pages through it on scroll. Continuation is **cursor-based**: each
    window returns an opaque ``cursor`` that the next request passes back, so a
    sequential scroll never re-scans the rows it already read. Pass ``cursor=
    None`` for the first window.

    For CSV the cursor is a byte offset, so seeking to the next window is O(1)
    and each fetch costs only O(limit) — no matter how deep into the file you
    are. (Streaming ``.xlsx`` can't be byte-seeked, so its cursor is the next
    row index and openpyxl still scans from the top; that's inherent to the
    format, but the API is uniform.)

    Returns ``{"rows": [...], "eof": bool, "cursor": str | None}``. ``eof`` is
    True (and ``cursor`` None) once the end of the file is reached, so the caller
    stops — no full-file row count is ever needed.
    """
    limit = max(1, limit)
    if kind == UploadedFile.Kind.EXCEL:
        return _window_excel(source, columns, cursor, limit)
    return _window_csv(source, columns, cursor, limit)


def _decoded_lines(fh):
    """Yield decoded lines one ``readline()`` at a time.

    Feeding csv this way (instead of iterating the file) keeps the file's byte
    position exact: csv pulls only the physical lines it needs for each record
    — including quoted fields that span newlines — and never reads ahead, so
    ``fh.tell()`` after N records lands precisely on the next record boundary.
    """
    while True:
        raw = fh.readline()
        if not raw:
            return
        yield raw.decode("utf-8", errors="replace")


def _window_csv(source, columns: list[str], cursor: str | None, limit: int) -> dict:
    rows: list[dict] = []
    eof = False
    with _as_binary(source) as fh:
        if cursor:
            fh.seek(int(cursor))
        else:
            fh.readline()  # skip the header row

        reader = csv.reader(_decoded_lines(fh))
        for _ in range(limit):
            try:
                raw = next(reader)
            except StopIteration:
                eof = True
                break
            rows.append(_row_to_dict(columns, raw))

        next_cursor = fh.tell()  # a clean record boundary
        # We stopped at exactly `limit` records — peek one byte to see whether
        # more remain, without consuming the next record (next_cursor is saved).
        if not eof and not fh.read(1):
            eof = True

    return {"rows": rows, "eof": eof, "cursor": None if eof else str(next_cursor)}


def _window_excel(source, columns: list[str], cursor: str | None, limit: int) -> dict:
    from openpyxl import load_workbook

    # cursor is the 1-based worksheet row to resume at; None -> first data row
    # (worksheet row 2, since row 1 is the header).
    start = int(cursor) if cursor else 2

    with _as_binary(source) as fh:
        wb = load_workbook(filename=fh, read_only=True, data_only=True)
        try:
            ws = wb.active
            rows: list[dict] = []
            eof = True
            next_row = start
            for row in ws.iter_rows(min_row=start, values_only=True):
                if len(rows) >= limit:
                    eof = False
                    break
                values = ["" if v is None else str(v) for v in row]
                rows.append(_row_to_dict(columns, values))
                next_row += 1
            return {"rows": rows, "eof": eof, "cursor": None if eof else str(next_row)}
        finally:
            wb.close()


def sample_values(
    preview_rows: list[dict],
    columns: list[str],
    per_column: int,
    max_len: int,
) -> dict[str, list[str]]:
    """Distinct, non-empty sample values for each of ``columns``.

    Drawn from the preview rows already captured at upload time — this never
    touches the file. Feeds the LLM real examples of a column's contents so it
    matches the data's actual case/format.

    Rows are visited **spread across** the preview window rather than top-down,
    so the samples reflect more of the column's variety instead of clustering on
    the first few (often identical) rows. Values are de-duplicated, capped at
    ``per_column`` values and ``max_len`` characters each. The traversal order is
    deterministic, so the cache signature stays stable. A column with no
    non-empty preview values is omitted rather than mapped to an empty list.
    """
    if per_column <= 0:
        return {}

    n = len(preview_rows)
    # Spread-first row order: every `stride`-th row (covering the whole window),
    # then the rows in between. Distinct values are collected in this order, so a
    # column with >= per_column distinct values yields ones sampled across the
    # window; dedup still tops up from the in-between rows when needed.
    stride = max(1, n // per_column)
    order = [i for start in range(stride) for i in range(start, n, stride)]

    out: dict[str, list[str]] = {}
    for col in columns:
        seen: list[str] = []
        for idx in order:
            raw = preview_rows[idx].get(col)
            if raw is None:
                continue
            value = str(raw).strip()[:max_len]
            if not value or value in seen:
                continue
            seen.append(value)
            if len(seen) >= per_column:
                break
        if seen:
            out[col] = seen
    return out


def _row_to_dict(columns: list[str], values: list) -> dict:
    out: dict[str, str] = {}
    for idx, col in enumerate(columns):
        out[col] = str(values[idx]) if idx < len(values) else ""
    return out
