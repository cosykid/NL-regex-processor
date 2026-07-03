"""Distributed pattern-matching & replacement engine (PySpark).

The engine reads the uploaded file into a Spark DataFrame, evaluates a set of
per-column regex predicates (combined with AND/OR) to select rows, applies the
optional ``regexp_replace`` to matched rows, and writes the result back as
Parquet for paged reading.

Design notes
------------
* **Native column expressions, not row UDFs.** ``regexp_replace`` / ``rlike``
  run inside the JVM across partitions — no per-row Python round-trips — so the
  job scales horizontally as row count grows into the millions. A multi-column
  condition is just a combination of native ``rlike`` column expressions, so it
  scales the same way.
* **Single parse, cached.** The source is read/parsed once and cached; the
  row-count pass materialises it and the write reads from the cached frame, so
  the CSV is never re-parsed per action. Partitioning is left to Spark's native
  file splitting (a large CSV is split by size across tasks); we don't force a
  full shuffle via ``repartition`` — it dominated cost on small files and was
  redundant on large ones that are already split. NB: size-based splitting only
  applies when ``SPARK_CSV_MULTILINE`` is False; with it True (the default) the
  CSV is non-splittable and the read runs single-core.
* **Progress.** A background thread polls ``SparkContext.statusTracker()`` for
  task-completion fraction during each action and maps it onto a progress band
  surfaced through the polling API. Because small/local stages are a single task
  (no granularity) and Spark reports nothing during boot/commit gaps, the poller
  blends the real fraction with a time-based ease so the bar keeps moving
  instead of freezing at a band boundary and then jumping. See ``_ProgressPoller``.
* **Cancellation.** The same thread checks a cancel callback and calls
  ``cancelAllJobs()``, which aborts the in-flight Spark action.

``pyspark`` is imported lazily so the rest of the Django app (and the
pure-Python tests) import without a Spark/JVM install present.
"""
from __future__ import annotations

import csv
import logging
import threading
import time
from dataclasses import dataclass, field
from functools import reduce
from typing import Callable

from django.conf import settings

from . import storage
from .exceptions import JobCancelled, SparkProcessingError
from .regex_safety import escape_replacement

logger = logging.getLogger("processing")

ProgressCb = Callable[[int, str], None]
CancelCb = Callable[[], bool]

# Internal boolean column appended to the written result: True for every row
# whose per-column predicates satisfy the combinator (AND/OR) — regardless of
# whether the replacement actually changed the text. Lets the UI emphasise affected
# rows, offer an affected-only view, and scope an export. Prefixed/suffixed to
# avoid colliding with a real column from the source file; stripped from every
# user-facing column list and export.
MATCH_FLAG_COLUMN = "__nlrx_matched__"

# Concrete output actions the engine understands (mirrors jobs.models.Job.Action
# minus `auto`, which is resolved to one of these before Spark ever runs).
# `find` is neither: it writes the data through unchanged — only the match flag
# column and the match counts carry the result.
CELL_ACTIONS = frozenset({"replace", "mask", "extract"})
ROW_ACTIONS = frozenset({"keep", "drop"})

# What a mask writes over the matched text when the request names no mask string.
# A run of bullets reads as a redaction in the grid without leaking the length or
# content of the original value.
DEFAULT_MASK_TOKEN = "••••"


@dataclass
class ReplacementStats:
    total_rows: int = 0
    matched_rows: int = 0
    columns: list[str] = field(default_factory=list)
    result_path: str = ""


def build_match_condition(F, predicates: list[dict], combinator: str):
    """Combine per-column ``rlike`` tests into one row-match column expression.

    ``all`` -> every predicate must hold (AND); ``any`` -> at least one (OR).
    Kept as a standalone function (taking the Spark ``functions`` module ``F``)
    so the AND/OR logic can be unit-tested without a running JVM.

    Each column is null-coalesced to ``""`` before ``rlike``: Spark reads an
    empty CSV field as SQL NULL, and ``rlike`` on NULL is NULL (never true), so
    an "is empty/blank" pattern like ``^\\s*$`` would match nothing. Coalescing
    makes a missing cell read as an empty string — exactly what the user means
    by "null/blank". It's a no-op for ordinary patterns: a NULL never satisfies
    them either way.
    """
    conds = [
        F.coalesce(F.col(p["column"]), F.lit("")).rlike(p["pattern"])
        for p in predicates
    ]
    join = (lambda a, b: a & b) if combinator == "all" else (lambda a, b: a | b)
    return reduce(join, conds)


def cell_action_expr(F, column: str, pattern: str, action: str,
                     replacement: str, mask_token: str):
    """Rewritten value of one predicate column under a cell action.

    Only fires inside matched rows (guarded by ``MATCH_FLAG_COLUMN``), leaving
    every other row untouched:

    * ``replace`` — swap the matched text for ``replacement`` (blank = delete).
    * ``mask``    — overwrite the matched text with ``mask_token``.
    * ``extract`` — collapse the cell to just the matched text; additionally
      guarded by this column's own ``rlike`` so an OR row never blanks a column
      that had no match of its own.

    Both ``replacement`` and ``mask_token`` must already be escaped for Spark's
    ``regexp_replace`` (see :func:`regex_safety.escape_replacement`). Taking ``F``
    keeps the branch logic testable with a symbolic stand-in, no JVM required.

    The matched cell is null-coalesced to ``""`` before rewriting: Spark reads an
    empty CSV field as SQL NULL, and ``regexp_replace``/``regexp_extract`` return
    NULL on NULL input — so "replace blank cells with 0" would silently no-op on
    the very cells it targets. Coalescing lets an empty/blank pattern fire on a
    missing cell; ``.otherwise(c)`` still returns the untouched original for every
    non-matched row.
    """
    c = F.col(column)
    cc = F.coalesce(c, F.lit(""))
    if action == "extract":
        return F.when(
            F.col(MATCH_FLAG_COLUMN) & cc.rlike(pattern),
            F.regexp_extract(cc, pattern, 0),
        ).otherwise(c)
    sub = mask_token if action == "mask" else replacement
    return F.when(
        F.col(MATCH_FLAG_COLUMN),
        F.regexp_replace(cc, pattern, sub),
    ).otherwise(c)


def get_spark(job_id: str):
    """Get/create the SparkSession for this worker process."""
    import os

    from pyspark.sql import SparkSession  # type: ignore[import-not-found]

    # Driver heap must be handed to the JVM *at launch*: a SparkSession.config()
    # set is ignored once the in-process (``local[*]``) JVM is already up, so the
    # launcher only honours the ``SPARK_DRIVER_MEMORY`` env var. Set it before the
    # first getOrCreate (this call, warm-up included) so the value actually takes.
    if settings.SPARK_DRIVER_MEMORY:
        os.environ.setdefault("SPARK_DRIVER_MEMORY", settings.SPARK_DRIVER_MEMORY)

    builder = (
        SparkSession.builder.appName(f"{settings.SPARK_APP_NAME}-{job_id}")
        .master(settings.SPARK_MASTER_URL)
        .config("spark.sql.shuffle.partitions", settings.SPARK_SHUFFLE_PARTITIONS)
        .config("spark.ui.showConsoleProgress", "false")
    )
    # S3A filesystem + credential-provider config when the storage backend is
    # S3 (empty dict for local, so this is a no-op there).
    for key, value in storage.spark_hadoop_conf().items():
        builder = builder.config(key, value)

    driver_host = os.environ.get("SPARK_DRIVER_HOST")
    if driver_host:
        builder = builder.config("spark.driver.host", driver_host)
    return builder.getOrCreate()


def warm_spark() -> None:
    """Boot the SparkSession/JVM ahead of the first job.

    Called from the Celery ``worker_process_init`` signal so each prefork child
    pays the ~10-15s cold JVM start at boot instead of on the user's first
    "apply pattern" (which otherwise freezes the progress bar mid-run while the
    JVM comes up). A trivial action forces full initialisation — JVM, SQL
    engine, and codegen — so the first real job is fully warm. Best-effort: any
    failure (no JVM, cluster unreachable) is logged and swallowed so the worker
    still starts and jobs fall back to lazy session creation.
    """
    try:
        spark = get_spark("warmup")
        spark.range(1).count()
        logger.info("Spark session warmed up")
    except Exception as exc:  # noqa: BLE001 - warm-up is best-effort
        logger.warning("Spark warm-up skipped: %s", exc)


def _prepare_input(input_locator: str, kind: str) -> str:
    """Resolve the Spark-readable URI for the source file.

    CSV is read in place — a local path or an ``s3a://`` URI, straight from the
    storage backend. Excel isn't Spark-native, so it's localized (a no-op for
    the local backend, a download for S3) and streamed to a temp CSV that Spark
    reads from the local filesystem. Excel is inherently bounded (~1M rows), so
    streaming it with openpyxl's read-only iterator is cheap; the million-row
    distributed scaling target is the CSV path.

    NOTE: the Excel branch writes a *local* temp CSV, which the default
    ``local[*]`` runtime reads directly. Under a standalone cluster
    (``spark://``) remote executors wouldn't see it — Excel there would need the
    converted CSV staged back to shared storage first.
    """
    from jobs.models import UploadedFile

    if kind != UploadedFile.Kind.EXCEL:
        return storage.spark_read_uri(input_locator)

    from openpyxl import load_workbook

    src = storage.localize(input_locator, suffix=".xlsx")
    out = src.with_suffix(src.suffix + ".converted.csv")
    if out.exists():
        return str(out)

    wb = load_workbook(filename=str(src), read_only=True, data_only=True)
    try:
        ws = wb.active
        if ws is None:
            raise SparkProcessingError(
                f"Excel file has no active worksheet: {input_locator}"
            )
        with open(out, "w", newline="", encoding="utf-8") as fh:
            writer = csv.writer(fh)
            for row in ws.iter_rows(values_only=True):
                writer.writerow(["" if v is None else v for v in row])
    finally:
        wb.close()
    return str(out)


class _ProgressPoller(threading.Thread):
    """Polls Spark task progress and the cancel flag during an action.

    Reported progress is the *max* of two signals so the bar always advances:

    * **Real task fraction** — ``completed / total`` tasks. Only *completed*
      tasks count; in-flight tasks are ignored here. On small/local data a stage
      is a single task, so crediting the one running task (fully, or even
      partially) would jump the bar a big chunk of the band the instant it went
      active — the flip side of the old "freeze then jump to 95%" symptom. A
      multi-task stage still gets real granularity as tasks finish.
    * **Time-based ease** — a synthetic fraction ``t / (t + TAU)`` that carries
      the bar forward when Spark reports no task-level granularity: a single
      in-flight task, or the dead gaps during JVM/session boot and the Parquet
      write's commit phase, which previously froze the bar at a band boundary.
      Capped below 1 so it approaches but never *reaches* the band end — only
      real task completion closes that final gap.

    Emissions are monotonic within the band: the bar never steps backward and we
    skip redundant same-value updates.
    """

    # Time constant (s) for the synthetic ease; at t=TAU it sits at half the
    # remaining band. Tuned so a multi-second stage visibly moves.
    _EASE_TAU = 6.0
    # Ceiling for the time-only ease so it can't fill the band on its own.
    _EASE_CAP = 0.9

    def __init__(self, spark, band: tuple[int, int], stage_label: str,
                 progress_cb: ProgressCb, cancel_cb: CancelCb):
        super().__init__(daemon=True)
        self._sc = spark.sparkContext
        self._start, self._end = band
        self._label = stage_label
        self._progress_cb = progress_cb
        self._cancel_cb = cancel_cb
        # NB: not `self._stop` — that name shadows Thread._stop(), the private
        # method CPython calls during join(), which would blow up with
        # "'Event' object is not callable".
        self._stop_event = threading.Event()
        self._last_pct = self._start  # monotonic floor for emitted progress
        self._t0 = 0.0
        self.cancelled = False

    def run(self) -> None:  # pragma: no cover - timing/JVM dependent
        tracker = self._sc.statusTracker()
        self._t0 = time.monotonic()
        while not self._stop_event.wait(0.5):
            if self._cancel_cb():
                self.cancelled = True
                self._sc.cancelAllJobs()
                return
            try:
                total = completed = 0
                for sid in tracker.getActiveStageIds():
                    info = tracker.getStageInfo(sid)
                    if info:
                        total += info.numTasks
                        completed += info.numCompletedTasks
                real = completed / total if total else 0.0
                elapsed = time.monotonic() - self._t0
                eased = min(elapsed / (elapsed + self._EASE_TAU), self._EASE_CAP)
                frac = min(max(real, eased), 1.0)
                pct = int(self._start + frac * (self._end - self._start))
                if pct > self._last_pct:
                    self._last_pct = pct
                    self._progress_cb(pct, self._label)
            except Exception:  # noqa: BLE001 - progress is best-effort
                pass

    def stop(self) -> None:
        self._stop_event.set()


def _run_action(spark, band, label, progress_cb, cancel_cb, action):
    """Run a Spark action with live progress polling + cancellation."""
    poller = _ProgressPoller(spark, band, label, progress_cb, cancel_cb)
    poller.start()
    try:
        return action()
    except Exception as exc:  # noqa: BLE001
        if poller.cancelled or cancel_cb():
            raise JobCancelled("Job cancelled during Spark processing") from exc
        raise
    finally:
        poller.stop()
        poller.join(timeout=2)
        if poller.cancelled:
            raise JobCancelled("Job cancelled during Spark processing")


def run_replacement(
    *,
    job_id: str,
    input_locator: str,
    kind: str,
    predicates: list[dict],
    combinator: str,
    action: str,
    replacement_value: str,
    result_locator: str,
    progress_cb: ProgressCb,
    cancel_cb: CancelCb,
) -> ReplacementStats:
    """Select rows via a set of per-column predicates and write Parquet results.

    Each predicate ``{"column", "pattern"}`` tests one column with ``rlike``;
    the tests are combined with ``combinator`` (``all`` = AND, ``any`` = OR) to
    decide whether a row matches.

    ``action`` then decides what to do with the selection, *only ever touching
    matched rows*:

    * ``replace`` / ``mask`` / ``extract`` — rewrite each predicate's column with
      its own pattern (see :func:`cell_action_expr`); every row is kept.
    * ``keep`` / ``drop`` — filter the dataset to the matched rows (or their
      complement); no cell is edited.
    * ``find`` — report only: no cell is edited and no row is filtered. The
      data passes through unchanged; the match flag column and the row counts
      are the entire result.

    So a compound condition like "name starts with A AND phone starts with 0"
    never edits or drops rows that don't satisfy the whole condition.
    """
    from pyspark.sql import functions as F  # type: ignore[import-not-found]
    from pyspark.storagelevel import StorageLevel  # type: ignore[import-not-found]

    # Set a stage *before* get_spark(): a cold session (warm-up disabled or
    # cluster mode) blocks here booting the JVM, and without this the bar would
    # sit frozen on the previous "regex ready" stage for the whole boot.
    progress_cb(14, "starting Spark engine")
    spark = get_spark(job_id)
    progress_cb(15, "reading file into Spark")

    read_uri = _prepare_input(input_locator, kind)
    # ``multiLine`` trades read parallelism for embedded-newline correctness: True
    # (default) parses quoted fields that span newlines but reads the file on a
    # single core; False restores size-based file splitting (all cores) for data
    # known to have no in-field newlines. See ``SPARK_CSV_MULTILINE``.
    df = (
        spark.read.option("header", True)
        .option("multiLine", settings.SPARK_CSV_MULTILINE)
        .option("escape", '"')
        .csv(read_uri)
    )
    all_columns = df.columns
    predicate_columns = [p["column"] for p in predicates]
    missing = [c for c in predicate_columns if c not in all_columns]
    if missing:
        raise SparkProcessingError(
            f"Target column(s) not found in file: {', '.join(missing)}"
        )

    if cancel_cb():
        raise JobCancelled("Job cancelled before processing")

    # A row matches when its per-column predicates combine to True (AND / OR).
    match_condition = build_match_condition(F, predicates, combinator)

    # Parse the source ONCE. Caching lets the count pass below materialise the
    # frame so the later write reads from memory instead of re-reading and
    # re-parsing the CSV. (MEMORY_AND_DISK spills gracefully when the data is
    # larger than memory.)
    df = df.persist(StorageLevel.MEMORY_AND_DISK)

    # Total rows AND matched rows in a single pass — previously two separate
    # ``.count()`` actions, each re-scanning the whole file. This action also
    # materialises the cache the write then reuses.
    counts = _run_action(
        spark, (15, 42), "scanning rows", progress_cb, cancel_cb,
        lambda: df.agg(
            F.count(F.lit(1)).alias("total"),
            F.sum(F.when(match_condition, 1).otherwise(0)).alias("matched"),
        ).first(),
    )
    total_rows = int(counts["total"] or 0)
    matched_rows = int(counts["matched"] or 0)

    # Tag each row with whether it matched, computed from the ORIGINAL values
    # (before any rewrite touches the target columns). Adding this column first
    # means the later projections leave it intact — and lets a cell action be
    # scoped to matched rows (and a row action to filter) by referring to the flag.
    transformed = df.withColumn(MATCH_FLAG_COLUMN, match_condition)

    if action in ROW_ACTIONS:
        # Row actions don't edit cells — they filter the dataset to the matched
        # rows (`keep`) or their complement (`drop`). The flag column rides along
        # so the result schema is uniform (results.py strips it either way).
        keep = F.col(MATCH_FLAG_COLUMN) if action == "keep" else ~F.col(MATCH_FLAG_COLUMN)
        transformed = transformed.filter(keep)
    elif action in CELL_ACTIONS:
        # Cell actions rewrite each predicate's column with its own pattern, but
        # only where the whole row matched, leaving every other row (and every
        # non-predicate column) untouched.
        replacement = escape_replacement(replacement_value)
        mask_token = escape_replacement(replacement_value or DEFAULT_MASK_TOKEN)
        for p in predicates:
            col = p["column"]
            transformed = transformed.withColumn(
                col,
                cell_action_expr(
                    F, col, p["pattern"], action, replacement, mask_token
                ),
            )

    write_uri = storage.spark_write_uri(result_locator)
    _run_action(
        spark, (42, 95), f"applying {action} (Spark write)", progress_cb, cancel_cb,
        lambda: transformed.write.mode("overwrite").parquet(write_uri),
    )

    # Free the cached frame; the SparkSession is reused across jobs in a worker.
    df.unpersist()

    progress_cb(98, "finalising")
    return ReplacementStats(
        total_rows=total_rows,
        matched_rows=matched_rows,
        columns=all_columns,
        result_path=result_locator,
    )
