"""Exports over Parquet results — CSV and Excel (`processing.results`).

These exercise the reader/writer path directly (no Spark, no DB): a small
Parquet file with the internal match-flag column stands in for a real result.
"""
import duckdb
import pytest
from openpyxl import load_workbook

from processing import results
from processing.spark_engine import MATCH_FLAG_COLUMN


@pytest.fixture
def result_dir(tmp_path):
    """A one-file Parquet 'result' with three display columns + the match flag.

    Row values are deliberately awkward: a NULL, a control character Excel
    rejects, and an over-long string, so the coercion path gets tested.
    """
    d = tmp_path / "result"
    d.mkdir()
    con = duckdb.connect()
    con.execute(
        f"""
        COPY (
          SELECT * FROM (VALUES
            ('a\x07b', 42, TRUE),
            (NULL, 7, FALSE),
            (repeat('x', 40000), 1, TRUE)
          ) t(name, num, "{MATCH_FLAG_COLUMN}")
        ) TO '{d}/part.parquet' (FORMAT PARQUET)
        """
    )
    con.close()
    return str(d)


def _load(path):
    wb = load_workbook(path)
    ws = wb.active
    return ws, [[c.value for c in row] for row in ws.iter_rows()]


def test_write_xlsx_drops_flag_and_keeps_all_rows(result_dir):
    path = results.write_xlsx(result_dir)
    ws, rows = _load(path)

    # Header excludes the internal match-flag column.
    assert rows[0] == ["name", "num"]
    assert MATCH_FLAG_COLUMN not in rows[0]
    # Header + three data rows.
    assert ws.max_row == 4


def test_write_xlsx_sanitises_cells(result_dir):
    _, rows = _load(results.write_xlsx(result_dir))

    # Control char stripped, numeric type preserved, NULL stays empty.
    assert rows[1] == ["ab", 42]
    assert rows[2] == [None, 7]
    # Over-long string clamped to Excel's per-cell ceiling.
    assert len(rows[3][0]) == results.XLSX_MAX_CELL_LEN


def test_write_xlsx_matched_only(result_dir):
    _, rows = _load(results.write_xlsx(result_dir, matched_only=True))

    # Header + the two flagged rows only (the FALSE row is dropped).
    assert len(rows) == 3
    assert [r[1] for r in rows[1:]] == [42, 1]


def test_write_xlsx_rejects_too_many_rows(result_dir, monkeypatch):
    # Shrink the ceiling instead of materialising a million rows.
    monkeypatch.setattr(results, "XLSX_MAX_ROWS", 3)  # 2 data rows allowed; we have 3
    with pytest.raises(results.ResultTooLargeForExcel, match="rows"):
        results.write_xlsx(result_dir)


def test_write_xlsx_rejects_too_many_columns(result_dir, monkeypatch):
    # Shrink the ceiling instead of materialising 16k+ columns; the fixture has
    # two display columns.
    monkeypatch.setattr(results, "XLSX_MAX_COLS", 1)
    with pytest.raises(results.ResultTooLargeForExcel, match="columns"):
        results.write_xlsx(result_dir)
